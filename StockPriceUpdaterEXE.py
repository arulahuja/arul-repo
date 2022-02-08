#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jun 21 14:54:29 2021

@author: arulahuja
"""

import tkinter as tk
from tkinter import filedialog, Text
import os
import yfinance as yf
import openpyxl



def Addfile():
    filename = filedialog.askopenfilename(title = "Select Excel File")
    label = tk.Label(frame, text = filename+"------- SUCCESFULLY LOADED", bg="gray")
    label.pack()
    path = filename
    print(path)


    workbook = openpyxl.load_workbook(path)
    ticker_names = []
    
    
    for sheets in workbook.worksheets:
        portfolio = sheets
        i = 2
        for row in portfolio.iter_rows(min_row = 2):
            ticker_name = row[1].value
            j = 4
            for col in portfolio.iter_cols(min_col = 4, max_row=1):
                if col[0].value != None:
                    try:
                        price = yf.Ticker(ticker_name).history(start=col[0].value)["Close"].tolist()
                        portfolio.cell(row = i, column = j, value = price[0])
                    except:
                        portfolio.cell(row = i, column = j, value = "ERROR")
    
                else:
                    pass
                j = j+1
            i = i+1
    workbook.save(path)
    popup = tk.messagebox.showinfo("Succesul", "Stock Prices Updated")    



root = tk.Tk()
root.title("Stock Price Updater")
canvas = tk.Canvas(root, height=400, width=600, bg="#263D42")
canvas.pack()

frame = tk.Frame(root, bg="white")
frame.place(relwidth=0.8, relheight=0.3, anchor="c", relx=.5, rely=.5)

Openfile = tk.Button(root, text="Open File and Run", command=Addfile)
Openfile.pack()

root.mainloop()