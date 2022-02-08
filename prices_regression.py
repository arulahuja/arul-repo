#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Aug 24 20:03:18 2021

@author: arulahuja
"""

import openpyxl
from collections import defaultdict
import numpy as np
from sklearn import linear_model

path = "/Users/arulahuja/Desktop/Pebble Pricing/reportsdishwise 2/newtest.xlsx"
workbook = openpyxl.load_workbook(path, read_only=True)
flagger = False
pq = defaultdict(list)

for sheets in workbook.worksheets:
    i = 7
    for row in sheets.iter_rows(min_row = 1):
        if flagger is True:
            try:
                if "H/Hrs" in row[2].value:
                    if row[2].value not in pq.keys():
                        pq[row[2].value.split("H/Hrs ")[1]].append((row[8].value, row[9].value/row[8].value))
                    else:
                        pq[row[2].value.split("H/Hrs ")[1]].append((row[8].value, row[9].value/row[8].value))
                elif row[2].value in pq.keys():
                    pq[row[2].value].append((row[8].value, row[9].value/row[8].value))
                else:
                    pq[row[2].value].append((row[8].value, row[9].value/row[8].value))
            except:
                pass
            
            
            
            
        if row[0].value == "Liquidarity":
            flagger = True 
        if row[0].value is None:
            flagger = False
        i = i+1



# sheets = workbook["Sheet1"]          
        
# for row in sheets.iter_rows(min_row = 7):
#     if flagger is True:
#         if "H/Hrs" in row[2].value:
#             pq[row[2].value.split("H/Hrs ")] = (row[8].value, row[9].value/row[8].value)
#         else:
#             pq[row[2].value] = (row[8].value, row[9].value/row[8].value)
#     if row[0].value == "Liquidarity":
#         flag = True 
#     if row[0].value is None:
#         flagger = False





# REGRESSION

obj = pq["Absolut"]
unzipped = zip(*obj)
unzipped_list = list(unzipped)
q = unzipped_list[0]
p = unzipped_list[1]
p_array = np.array(p).reshape((-1, 1))
q_array = np.array(q)


# model = LinearRegression().fit(p_array, q_array)


# r_sq = model.score(p_array, q_array)
# print('coefficient of determination:', r_sq)
# print('slope:', model.coef_)
reg = linear_model.Ridge(alpha=.5)
reg.fit(p_array, q_array)
print(reg.coef_)
print(reg.intercept_)

